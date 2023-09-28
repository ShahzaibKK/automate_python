#! python3
# multiply_table.py - by Shahzaib kk - in Excel

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
import sys

wb = Workbook()
sheet: Worksheet = wb.active
if len(sys.argv) > 1:
    for row in range(1, int(sys.argv[1]) + 1):
        for col in range(1, int(sys.argv[1]) + 1):
            sheet.cell(row=row, column=col, value=row * col).font = Font(bold=True)
        # sheet.cell(row=i + 1, column=1, value=i).font = Font(bold=True)
        # sheet.cell(row=1, column=i + 1, value=i).font = Font(bold=True)

    wb.save("./Multitable.xlsx")
else:
    print("Plese proive integer from commond line")
