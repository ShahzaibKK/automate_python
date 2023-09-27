#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

wb = load_workbook("produceSales.xlsx")
sheet: Worksheet = wb.active

# The produce types and thier updated prices

PRICE_UPDATES = {"Garlic": 3.07, "Celery": 1.19, "Lemon": 1.27}

# TODO Loop through the row and update the prices.

for row_num in range(2, sheet.max_row):
    produce_name = sheet.cell(row=row_num, column=1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]


wb.save("updatedProduceSales.xlsx")
