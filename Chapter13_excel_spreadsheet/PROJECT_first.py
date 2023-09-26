#! python3
#  readCensusExcel.py - Tablates populatin and number of census tracts for
# each Country.

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import pprint

print("Opening workbook....")
wb: Workbook = load_workbook("censuspopdata.xlsx")
sheet: Worksheet = wb.active
country_data = {}

# TODO: Fill in countryData with each country's population and tracts.
print("Reading Row...")
for row in range(2, sheet.max_row + 1):
    state = sheet["B" + str(row)].value
    county = sheet["C" + str(row)].value
    pop = sheet["D" + str(row)].value

    country_data.setdefault(state, {})
    country_data[state].setdefault(county, {"tracts": 0, "pop": 0})
    country_data[state][county]["tracts"] += 1
    country_data[state][county]["pop"] += int(pop)


# TODO: Open a new text file and write the contents of countyData to it.
print("Waiting for results...")
with open("census2010.py", "w") as result_file:
    result_file.write("allData = " + pprint.pformat(country_data))
    print("DONE, Thanks")
