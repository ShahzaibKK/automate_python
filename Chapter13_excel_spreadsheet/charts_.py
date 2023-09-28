from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import BarChart, Reference

wb = Workbook()
sheet: Worksheet = wb.active

for i in range(1, 11):
    sheet.cell(row=i, column=1, value=i)

data = Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
chart = BarChart()
chart.title = "My Chart"
chart.add_data(data)

sheet.add_chart(chart, "C5")
wb.save("sampleChart.xlsx")
