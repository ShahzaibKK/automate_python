from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.fonts import Font
from random import random

wb: Workbook = Workbook()
print(wb.sheetnames)
sheet: Worksheet = wb.active
sheet.title = "Spam Bacon Eggs Sheet"
print(sheet.title)
print(wb.sheetnames)
wb.create_sheet(index=0)
wb.create_sheet(index=0, title="The King")
print(wb.sheetnames)
del wb["Sheet"]
del wb["The King"]
print(wb.sheetnames)
italic_24_font = Font(size=24, italic=True, bold=True)  # type: Font
for_total = Font(
    name="Century Gothic",
    size=24,
    italic=True,
    bold=True,
)  # type: Font
sheet.row_dimensions[1].height = 70
sheet["A1"] = "Hello World"
print(sheet["A1"].value)
for num in range(1, 9):
    sheet["B" + str(num)] = round(random(), 2) * 266
sheet["B9"].font = for_total
sheet["B9"] = "=Round(SUM(B1:B8),2)"
sheet.column_dimensions["B"].width = 20
sheet["A1"].font = italic_24_font
sheet.merge_cells("D5:F9")
sheet["D5"] = "King Khan"
sheet.freeze_panes = "C2"
wb.save("Test.xlsx")
