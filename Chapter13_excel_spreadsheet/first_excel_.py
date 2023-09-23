import openpyxl
from pathlib import Path

wb = openpyxl.load_workbook("example.xlsx")
print(type(wb))
print(Path.cwd())
print(wb.sheetnames)
sheet = wb["Sheet1"]
print(type(sheet))
print(sheet.title)
another_sheet = wb.active
print(another_sheet)
print(sheet["A1"].value)
b = sheet["B1"]
print(b.value)
print(f"Row {b.row}, Column {b.column} is {b.value} ")
print(f"Cell {b.coordinate} is {b.value}")
print(sheet["C1"].value)

print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

print(sheet.max_column, "\n", sheet.max_row)
