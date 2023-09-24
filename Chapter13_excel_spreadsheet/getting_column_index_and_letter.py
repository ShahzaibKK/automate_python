import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

wb = openpyxl.load_workbook("../example.xlsx")

sheet = wb["Sheet1"]
print(get_column_letter(1))
print(get_column_letter(27))
print(get_column_letter(sheet.max_column))

print(sheet.cell(row=1, column=1))
print(column_index_from_string("FU"))
print(get_column_letter(177))
# sheet.cell(row=2,column=2).value = "Babu lala"
print(sheet.cell(row=2,column=2).value)
