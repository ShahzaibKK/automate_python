import time
import re
import sys
import platform
import datetime
import uuid
import hashlib
import shutil
from pathlib import Path
from configparser import ConfigParser
from PIL import Image as PIL_Image, ImageDraw, ImageFont
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib.pagesizes import letter
from reportlab.platypus import (
    SimpleDocTemplate,
    Image,
    Table,
    TableStyle,
    PageBreak,
    Paragraph,
)
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
import logging
import requests

#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Stock PDF Generator

Author: Shahzaib KK +92 336 8311100

This Python script is designed to generate PDF documents from a collection of images.
It is specifically tailored to the needs of Khuram Tiles, Peshawar, for creating product catalogs.
The script compresses images, adds quantity information, and optionally includes a logo.
It is intended to automate the process of generating these catalogs.

Usage:
- Ensure you have the required libraries installed (Pillow, openpyxl, reportlab, requests).
- Prepare your images and Excel file as per the specified format.
- Run this script with the desired options to generate the PDF.

Options:
- 'logo': Include the company logo in the PDF.

Note:
This script is part of an automation project and is customized for a specific use case.
It may require adjustments for different scenarios.
"""

# Move logging configuration to the top so early logs are formatted correctly.
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s",
)

LICENSE_PATH = r"D:\KK's\automate_python\KINGs_work\Office_license.key"
CONFIG_PATH = r"D:\KK's\automate_python\KINGs_work\config.ini"
WATERMARK_TEXT = "King Khan Tiles Peshawar"


def verify_license():
    # Get the machine's MAC address
    mac_address = str(uuid.getnode())
    secret = "M_AMIR"

    # Hash the MAC address
    current_hashed_mac = hashlib.sha256((mac_address + secret).encode()).hexdigest()

    # Read the hashed MAC from the license file
    try:
        with open(LICENSE_PATH, "r") as license_file:
            stored_hashed_mac = license_file.read().strip()
    except FileNotFoundError:
        logging.error("Error: license.key not found.")
        sys.exit(1)
    except Exception as e:
        logging.error(f"Unexpected error while reading license file: {e}")
        sys.exit(1)

    # Validate the hash
    if current_hashed_mac != stored_hashed_mac:
        logging.error("Error: License validation failed. Unauthorized machine.")
        sys.exit(1)

    logging.info("License validated successfully.")


# Check for the config file existence
if not Path(CONFIG_PATH).exists():
    logging.error("Config file not found. Please ensure the path is correct.")
    sys.exit(1)

config = ConfigParser()
try:
    config.read(CONFIG_PATH)
except Exception as e:
    logging.error(f"Unexpected error while reading config file: {e}")
    sys.exit(1)

# Set up OS-specific paths (Windows only)
if platform.system() == "Windows":
    # For Windows, we use the Desktop folder
    desktop_path = Path.home() / "Desktop"

    ALL_RECORDS = desktop_path / "DATA"
    ALL_RECORDS.mkdir(exist_ok=True)

    COMPRESS_IMAGE_PATH = desktop_path / "DATA" / "compressed_images"
    COMPRESS_IMAGE_PATH.mkdir(exist_ok=True)

    try:
        all_pics = Path(config["Paths_bro"]["all_pics"])
        KT_LOGO = Path(config["Paths_bro"]["company_logo"])
    except KeyError as e:
        logging.error(f"Key '{e.args[0]}' is missing in config.ini.")
        sys.exit(1)
    destination_path_comp = COMPRESS_IMAGE_PATH
    formated_date = datetime.date.today().strftime("%d-%m-%Y")
    pdf_file = ALL_RECORDS / f"Available_Stock_{formated_date}.pdf"
    MISSING_FILES = set()
    MISSING_FILES_PATH = ALL_RECORDS / "Missing_Images.txt"
    AVAILABLE_STOCK = desktop_path / "available_stock.xlsx"
else:
    print("This script is intended for Windows operating systems.")
    sys.exit(1)


print("***** Create by Shahzaib KK +92 336 8311100 *****")
time.sleep(2)


def collect_articles():
    """Load Excel File and store the articles in a set."""
    wb = load_workbook(AVAILABLE_STOCK)
    if len(sys.argv) > 2:
        if sys.argv[2] == "DM":
            article_regex = re.compile(r"\d{2}DM\d{3}")
        else:
            article_regex = re.compile(
                r"^(?:[A-Z]+)?\d{2,3}(?:[A-Z]{2})?\d{2,6}", re.IGNORECASE
            )
    else:
        article_regex = re.compile(
            r"^(?:[A-Z]+)?\d{2,3}(?:[A-Z]{2})?\d{2,6}", re.IGNORECASE
        )

    sheet: Worksheet = wb.active  # Using the active worksheet

    articles_set = set()
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value:
                mo = article_regex.search(str(cell_value))
                if mo:
                    articles_set.add(mo.group())
    wb.close()
    return articles_set


def collect_qty():
    """
    Collects quantities from the Excel file.
    Returns a dictionary mapping article names to their quantities.
    """
    article_names = collect_articles()
    article_quantities = {}

    wb = load_workbook(AVAILABLE_STOCK)
    sheet: Worksheet = wb.active
    for article_name in article_names:
        article_regex = re.compile(rf"^{article_name}(\w+)?(\d+)?$")
        for row in sheet.iter_rows():
            for i, cell in enumerate(row):
                if cell.value:
                    mo = article_regex.search(str(cell.value))
                    if mo:
                        article_full_name = mo.group(0)
                        if i < len(row) - 1:
                            quantity_cell = row[i + 1]
                            if quantity_cell.value is not None:
                                article_quantities[article_full_name] = (
                                    quantity_cell.value
                                )
                            else:
                                # Force a '0' or some placeholder
                                article_quantities[article_full_name] = "Available"
    wb.close()
    return article_quantities


def check_compressed_files(path_and_name: Path):
    return path_and_name.exists()


def compress_images(
    image_path,
    destination_path,
    watermark_text,  # passed explicitly
    greater_than=0,
    target_resolution=(1800, 1800),
):
    size_MB = image_path.stat().st_size / (1024 * 1024)
    file_name = image_path.name
    compressed_file_name_dest = destination_path / f"compressed_{file_name}"

    if not check_compressed_files(compressed_file_name_dest):
        if size_MB > greater_than:
            image = PIL_Image.open(image_path)
            image = image.convert("RGB")
            image.thumbnail(target_resolution)

            # Add a watermark to the image
            draw = ImageDraw.Draw(image)
            font = ImageFont.truetype("arial.ttf", 45)
            text_bounding_box = font.getbbox(watermark_text)
            image_width, image_height = image.size
            text_width = text_bounding_box[2] - text_bounding_box[0]
            text_height = text_bounding_box[3] - text_bounding_box[1]
            text_position = (
                image_width - text_width - 15,
                image_height - text_height - 35,
            )

            draw.text(
                text_position,
                watermark_text,
                font=font,
                fill=(234, 215, 139, 128),
                anchor="la",
            )

            image.save(compressed_file_name_dest, format="JPEG", quality=60)
            logging.info(f"Compressed {file_name} to {compressed_file_name_dest}")
        else:
            shutil.copy(image_path, destination_path)
            logging.info(f"Copied {file_name} to {destination_path}")
    else:
        logging.info(f"{file_name} already compressed")
    return compressed_file_name_dest


def create_pdf(image_path: Path, output_pdf_path, logo_path=None):
    folder = image_path.parent
    qty = collect_qty()
    # Sort image files to prioritize those starting with "compressed_36DM"
    image_paths = sorted(
        folder.glob("*"), key=lambda path: not path.name.startswith("compressed_36DM")
    )

    doc = SimpleDocTemplate(output_pdf_path, pagesize=letter)
    doc.background = colors.black
    elements = []

    if logo_path:
        logo = Image(logo_path, width=7 * inch, height=8 * inch)
        elements.append(logo)

    categories = {
        "36": "12x24 Glaze",
        "M36": "12x24 Matt",
        "40": "16x16 Glaze",
        "25": "10x20 Glaze",
        "M30": "12x12 Matt",
        "M40": "16x16 Matt",
        "M25": "10x20 Matt",
        "60": "24x24 Glaze",
        "612": "24x48 Glaze",
    }
    current_category = None
    category_lookup = {key: category for key, category in categories.items()}

    for img_path in image_paths:
        pure_str = img_path.stem[11:]
        article_regex_pattern = rf"^{pure_str}(\w+)?(\d+)?$"
        article_regex = re.compile(article_regex_pattern)

        category_text = category_lookup.get(pure_str[:3])  # Try first 3 chars
        if not category_text:
            category_text = category_lookup.get(pure_str[:2])
        if category_text and current_category != category_text:
            current_category = category_text
            if elements:
                elements.append(PageBreak())
            category_paragraph = Paragraph(
                f"Category: {current_category}", getSampleStyleSheet()["Heading2"]
            )
            elements.append(category_paragraph)
            logging.info(f"Added new category: {current_category}")

        flowables = []
        pdf_image = Image(img_path, width=5.4 * inch, height=7.5 * inch)
        flowables.append(pdf_image)

        data = [["Article", "Quantity"]]
        for key in qty:
            mo = article_regex.search(key)
            if mo:
                article = key
                quantity = str(qty.get(mo.group()))
                data.append([article, quantity])

        table = Table(data)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 13),
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 1), (-1, -1), 13),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                    ("TOPPADDING", (0, 1), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("TEXTCOLOR", (0, 1), (-1, -1), colors.blue),
                    ("GRID", (0, 0), (-1, -1), 1.2, colors.black),
                ]
            )
        )
        flowables.append(table)
        elements.extend(flowables)
    doc.build(elements)


def get_world_time():
    """Fetch current time from worldtimeapi.org."""
    try:
        url = "http://worldtimeapi.org/api/timezone/Asia/Karachi"
        logging.debug(f"Fetching time from {url}")
        res = requests.get(url, timeout=10)
        res.raise_for_status()
        json_data = res.json()
        return datetime.datetime.fromisoformat(json_data["datetime"])
    except (requests.RequestException, KeyError, ValueError) as e:
        logging.error(f"Error fetching time from worldtimeapi.org: {e}")
        return None


def get_timeapi_time():
    """Fetch current time using TimeAPI."""
    try:
        timeapi_url = "https://timeapi.io/api/Time/current/zone?timeZone=Asia/Karachi"
        logging.debug(f"Fetching time from {timeapi_url}")
        res = requests.get(timeapi_url, timeout=10)
        res.raise_for_status()
        json_data = res.json()
        return datetime.datetime.fromisoformat(json_data["dateTime"])
    except (requests.RequestException, KeyError, ValueError) as e:
        logging.error(f"Error fetching time from TimeAPI: {e}")
        return None


def get_current_time():
    """Fetches the current time using APIs, with fallback to local time."""
    current_time = get_world_time()
    if not current_time:
        logging.error("worldtimeapi.org failed. Trying TimeAPI.")
        current_time = get_timeapi_time()
    if not current_time:
        logging.error("Both APIs failed. Falling back to local system time.")
        current_time = datetime.datetime.now(datetime.timezone.utc)
    return current_time


def validate_license():
    """Validates the program license based on expiration date."""
    current_time = get_current_time()
    logging.debug(f"Current time: {current_time}")
    EXPIRE = datetime.datetime(2025, 12, 31, 23, 59, 59, tzinfo=datetime.timezone.utc)
    if current_time.tzinfo is None:
        current_time = current_time.replace(tzinfo=datetime.timezone.utc)
    if EXPIRE < current_time:
        logging.error(
            "\n\n\n\n\n\n\tYour program license has expired. Please contact Shahzaib KK +92 336 8311100 to renew your license."
        )
        sys.exit()
    print("License is valid. Program continues...")


if __name__ == "__main__":
    verify_license()
    validate_license()

    # Determine watermark text based on command-line arguments.
    if len(sys.argv) > 1:
        if sys.argv[1] == "logo":
            watermark_text = WATERMARK_TEXT
        else:
            watermark_text = ""
    else:
        watermark_text = ""

    compress_images_path = None
    for article in collect_articles():
        file_name_ = article + ".jpg"
        article_path = Path.joinpath(all_pics, file_name_)
        if article_path.is_file():
            compress_images_path = compress_images(
                article_path,
                destination_path_comp,
                watermark_text,
                greater_than=0,
            )
        else:
            logging.error(f"Not Available: {article_path.stem}")
            MISSING_FILES.add(article_path)

    logging.info("\n\n\tGenerating PDF. Please wait...")

    with open(MISSING_FILES_PATH, "w") as missing_report:
        missing_report.writelines(
            f"{missing_file.stem}\n" for missing_file in MISSING_FILES
        )

    if compress_images_path:
        if len(sys.argv) > 1 and sys.argv[1] == "logo":
            create_pdf(compress_images_path, str(pdf_file), KT_LOGO)
            watermark_text = WATERMARK_TEXT
        else:
            create_pdf(compress_images_path, str(pdf_file))
            watermark_text = ""

    logging.info(f"\n\n\n\n\t ****PDF was Created****\n: {pdf_file}")
