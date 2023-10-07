import re, pprint

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Collect your article names using a more descriptive variable name
article_names = collect_articels()

# Load the Excel workbook
wb = load_workbook(r"D:\Khuram Tiles\Main Files\Huamei Ceramics\available_stock.xlsx")
sheet = wb["Table 1"]


# Initialize a dictionary to store article names and their quantities
article_quantities = {}

for article_name in article_names:
    article_regex = re.compile(rf"{article_name}(\w+)?(\d+)?")  # Adjusted regex

    for row in sheet.iter_rows():
        for i, cell in enumerate(row):
            if cell.value:
                mo = article_regex.search(str(cell.value))
                if mo:
                    article_full_name = mo.group(0)  # Get the complete article name
                    # Ensure the current cell is not the last in the row
                    if i < len(row) - 1:
                        quantity_cell = row[i + 1]
                        if quantity_cell.value:
                            article_quantities[article_full_name] = quantity_cell.value

    # Print the article names and their quantities
    pprint.pprint(article_quantities)

    # Close the workbook
    wb.close()
