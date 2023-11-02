#! python3
# Excel_to_CSV - Convert xsls to csv file by Shahzaib KK

import openpyxl, csv
from pathlib import Path

excel_files = Path(
    r"C:\Users\shahz\Downloads\automate_online-materials\excelSpreadsheets"
)
csv_files_path = Path.home() / "Desktop/CSV_Files"
if not csv_files_path.exists():
    csv_files_path.mkdir()


for excel_file in excel_files.iterdir():
    wb = openpyxl.load_workbook(excel_file)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        csv_file_name = f"{excel_file.stem}_{sheet.title}.csv"
        csv_file_path = csv_files_path / csv_file_name
        with csv_file_path.open("w", newline="") as csv_file_obj:
            csv_writer_obj = csv.writer(csv_file_obj)
            for row in sheet.iter_rows(values_only=True):
                csv_writer_obj.writerow(row)
